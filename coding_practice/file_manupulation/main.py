"""
Convert the full name from Input.txt file to first and last name and then input it in the output.csv
"""
import csv

with open(file="input_file.txt", mode="r") as file:
    names = file.readlines()

    # for name in names:
    #     first_name, last_name = name.split()
    #     print(f"First: {first_name}, Last: {last_name}")

# The newline='' parameter tells Python not to do this automatic conversion, letting the CSV module handle line endings properly.
with open(file="output.csv", mode="w", newline='') as csvfile:
    writer = csv.writer(csvfile)

    writer.writerow(['First Name', 'Last Name'])

    for name in names:
        name = name.strip()  # Remove \n and whitespace
        first_name, last_name = name.split()
        writer.writerow([first_name, last_name])
        print(f"First: {first_name}, Last: {last_name}")


"""
Ways to make headers bold and add an empty line in CSV files

from openpyxl import Workbook
from openpyxl.styles import Font

with open(file="input_file.txt", mode="r") as file:
    names = file.readlines()

# Create workbook
wb = Workbook()
ws = wb.active

# Write bold header
ws['A1'] = 'First Name'
ws['B1'] = 'Last Name'
ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)

# Empty row at row 2 (automatically there if you start writing at row 3)

# Write data starting from row 3
row = 3
for name in names:
    name = name.strip()
    if name:
        first_name, last_name = name.split()
        ws.cell(row=row, column=1, value=first_name)
        ws.cell(row=row, column=2, value=last_name)
        row += 1

wb.save('output.xlsx')
"""